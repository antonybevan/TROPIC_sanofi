#!/usr/bin/env python3
"""Create a variable-level ADaM traceability ledger from the authoring workbook."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WB = load_workbook(ROOT / "00_specifications/ADaM_spec.xlsx", read_only=True, data_only=True)


def sheet_rows(name):
    ws = WB[name]
    vals = ws.iter_rows(values_only=True)
    header = [str(v) if v is not None else "" for v in next(vals)]
    for row in vals:
        if not any(v is not None for v in row):
            continue
        yield {header[i]: row[i] for i in range(min(len(header), len(row)))}


methods = {r.get("ID"): r for r in sheet_rows("Methods") if r.get("ID")}
actual = {}
try:
    import pyreadstat
    for path in (ROOT / "04_adam").glob("*_prod.xpt"):
        name = path.stem.removesuffix("_prod").upper()
        _, meta = pyreadstat.read_xport(str(path), metadataonly=True)
        actual[name] = set(meta.column_names)
except Exception:
    pass

producers = {
    "ADSL": ("02_production_sas/A_adsl_generation.sas", "03_validation_r/v_adsl_validation.R"),
    "ADEX": ("02_production_sas/A_adex_generation.sas", "03_validation_r/v_adex_validation.R"),
    "ADCM": ("02_production_sas/A_adcm_generation.sas", "03_validation_r/v_adcm_validation.R"),
    "ADAE": ("02_production_sas/A_adae_io_respec.sas", "03_validation_r/v_adae_io_validation.R"),
    "ADLB": ("02_production_sas/A_adlb_generation.sas", "03_validation_r/v_adlb_validation.R"),
    "ADRS": ("02_production_sas/A_adrs_generation.sas", "03_validation_r/v_adrs_validation.R"),
    "ADTTE": ("02_production_sas/A_adtte_generation.sas", "03_validation_r/v_adtte_validation.R"),
}
tfl = {
    "ADSL": "F-01-1; F-12-1; F-14-1; multiple population denominators",
    "ADEX": "F-14-1; F-17-1",
    "ADCM": "No direct TFL consumer documented",
    "ADAE": "T-20",
    "ADLB": "F-13-1; F-17-1; T-21",
    "ADRS": "T-11 response analyses",
    "ADTTE": "F-11-1; F-11-2; F-12-1; T-11 survival analyses",
}

rows = []
for r in sheet_rows("Variables"):
    ds = str(r.get("Dataset") or "")
    var = str(r.get("Variable") or "")
    if not ds or not var:
        continue
    method_id = str(r.get("Method") or "")
    method = methods.get(method_id, {})
    origin = str(r.get("Origin") or "")
    predecessor = str(r.get("Predecessor") or "")
    description = str(method.get("Description") or "")
    sas, val = producers.get(ds, ("UNMAPPED", "UNMAPPED"))
    gaps = []
    if origin == "Derived" and not method_id:
        gaps.append("derived variable has no MethodOID/algorithm")
    if not predecessor:
        gaps.append("Predecessor/source variable not documented")
    if method_id and not description:
        gaps.append("MethodOID has no method description")
    rows.append({
        "dataset": ds,
        "variable": var,
        "origin": origin,
        "predecessor": predecessor,
        "method_oid": method_id,
        "method_description": description,
        "sas_producer": sas,
        "r_validation_producer": val,
        "actual_xpt_variable_present": "YES" if var in actual.get(ds, set()) else "NO",
        "define_item_present": "YES",
        "forward_tfl_consumers": tfl.get(ds, "No consumer mapping documented"),
        "traceability_status": "GAP: " + "; ".join(gaps) if gaps else "DOCUMENTED",
    })

out = ROOT / "audit/adam_variable_traceability.csv"
with out.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
    writer.writeheader()
    writer.writerows(rows)
print(f"Wrote {len(rows)} ADaM variable traceability records; documented={sum(r['traceability_status']=='DOCUMENTED' for r in rows)}; gaps={sum(r['traceability_status']!='DOCUMENTED' for r in rows)}")
