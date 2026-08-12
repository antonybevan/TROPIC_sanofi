"""Data-free regression controls for repaired analysis and metadata contracts."""

from __future__ import annotations

from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _text(relative: str) -> str:
    return (ROOT / relative).read_text(encoding="utf-8")


def test_ars_csv_writer_is_cross_platform_lf_deterministic() -> None:
    program = _text("platform/build_ars.py")
    assert 'lineterminator="\\n"' in program


def test_adtte_uses_typed_components_and_rejects_pre_origin_dates() -> None:
    r_program = _text("04_analysis_datasets/programs/r/v_adtte_validation.R")
    sas_program = _text("04_analysis_datasets/programs/sas/A_adtte_generation.sas")

    assert 'filter(PARAMCD == "OVRLRESP", AVALC == "PD")' in r_program
    assert 'filter(PARAMCD == "PSPROG", AVALC == "Y")' in r_program
    assert "Exploratory BSGRESP and generic CLINPROG are excluded" in r_program
    assert "d$ADT < d$STARTDT" in r_program
    assert "ADT < STARTDT" in sas_program


def test_sas_pfs_avoids_all_missing_min_operation_note() -> None:
    sas_program = _text("04_analysis_datasets/programs/sas/A_adtte_generation.sas")

    assert "nmiss(tumor_prog_dt, psa_prog_dt, _pain_dt, _death_dt) < 4" in sas_program
    assert "_event_dt = min(tumor_prog_dt, psa_prog_dt, _pain_dt, _death_dt)" in sas_program


def test_adsl_rejects_partial_iso_dates_before_last_alive_aggregation() -> None:
    r_program = _text("04_analysis_datasets/programs/r/v_adsl_validation.R")
    admiral_program = _text("04_analysis_datasets/programs/r/admiral_adsl.R")

    assert 'grepl("^\\\\d{4}-\\\\d{2}-\\\\d{2}$", value)' in r_program
    assert "ymd(substring(as.character(x), 1, 10), quiet = TRUE)" not in r_program
    assert 'grepl("^\\\\d{4}-\\\\d{2}-\\\\d{2}$", value)' in admiral_program
    assert "ymd(substr(as.character(x), 1, 10), quiet = TRUE)" not in admiral_program
    assert "RANDDT  = date10(RFSTDTC)" in admiral_program


def test_spec_data_gate_preserves_planned_actual_treatment_distinction() -> None:
    program = _text("04_analysis_datasets/programs/r/spec_data_checks.R")

    assert "TRT01P/TRT01PN planned-treatment mapping is inconsistent" in program
    assert "TRT01A/TRT01AN actual-treatment mapping is inconsistent" in program
    assert "planned/actual treatment discrepancy(ies) retained" in program
    assert "TRT01A is not equal to the DM-authoritative TRT01P arm" not in program


def test_ttsae_end_is_parameter_bounded_without_silent_flooring() -> None:
    r_program = _text("04_analysis_datasets/programs/r/v_adtte_validation.R")
    sas_program = _text("04_analysis_datasets/programs/sas/A_adtte_generation.sas")

    for token in (
        "coalesce(LSTALVDT, STUDY_CUTOFF_DT)",
        "TRTEDT + days(SAFETY_FOLLOWUP_DAYS)",
        "STUDY_CUTOFF_DT",
    ):
        assert token in r_program
    assert "trtedt + &SAFETY_FOLLOWUP_DAYS." in sas_program
    assert "max(_safety_end" not in sas_program.lower()
    assert "pmax(safety_end" not in r_program.lower()


def test_adex_uses_primary_iv_cycles_and_source_rdi() -> None:
    r_program = _text("04_analysis_datasets/programs/r/v_adex_validation.R")
    sas_program = _text("04_analysis_datasets/programs/sas/A_adex_generation.sas")

    for token in ("MITOX", "XRP", "CABAZ", "EXTRINT", "EXDOSE2 > 0"):
        assert token in r_program
        assert token.lower() in sas_program.lower()
    assert "n_distinct(VISITNUM[!is.na(EXDOSE2) & EXDOSE2 > 0])" in r_program
    assert "count(distinct case when exdose2 > 0 then visitnum" in sas_program.lower()


def test_occurrence_and_lab_outputs_preserve_source_sequence_keys() -> None:
    adcm_r = _text("04_analysis_datasets/programs/r/v_adcm_validation.R")
    adcm_sas = _text("04_analysis_datasets/programs/sas/A_adcm_generation.sas")
    adlb_r = _text("04_analysis_datasets/programs/r/v_adlb_validation.R")
    adlb_sas = _text("04_analysis_datasets/programs/sas/A_adlb_generation.sas")

    assert 'anyDuplicated(adcm[c("USUBJID", "CMSEQ")])' in adcm_r
    assert "by usubjid CMSEQ" in adcm_sas
    assert '"lbdy", "LBSEQ"' in adlb_r
    assert "by usubjid PARAMCD AVISITN lbdy LBSEQ" in adlb_sas
    assert "AVALC = LBSTRESC" in adlb_r
    assert "lb.lbstresc as AVALC" in adlb_sas


def test_tfl_gate_uses_controlled_alpha_and_stratified_methods() -> None:
    program = _text("05_outputs/tfl/tfl_generation.R")
    stats = _text("05_outputs/tfl/tfl_stats.R")

    assert "pfs_pval < FINAL_ALPHA" in program
    assert "os_pval < FINAL_ALPHA" in program
    assert "Stratified Cox HR" in program
    assert "Stratified log-rank" in program
    assert 'ties = "efron"' in stats
    assert "na.action = stats::na.fail" in stats


def test_define_has_complete_adtte_value_level_metadata() -> None:
    root = ET.parse(ROOT / "03_metadata/define/define.xml").getroot()
    ns = {
        "odm": "http://www.cdisc.org/ns/odm/v1.3",
        "def": "http://www.cdisc.org/ns/def/v2.1",
    }
    value_list = root.find(".//def:ValueListDef[@OID='VL.ADTTE.AVAL']", ns)
    assert value_list is not None
    refs = value_list.findall("odm:ItemRef", ns)
    assert len(refs) == 6

    expected = {"OS", "PFS", "TTPSA", "TTUMOR", "TTPAIN", "TTSAE"}
    where_values = set()
    for clause in root.findall(".//def:WhereClauseDef", ns):
        if clause.attrib.get("OID", "").startswith("WC.ADTTE.PARAMCD.EQ."):
            where_values.update(
                node.text for node in clause.findall(".//odm:CheckValue", ns)
            )
    assert expected <= where_values


def test_define_identifies_non_submission_context_and_valid_aeser_oid() -> None:
    ns = {
        "odm": "http://www.cdisc.org/ns/odm/v1.3",
        "def": "http://www.cdisc.org/ns/def/v2.1",
    }
    for relative in ("03_metadata/define/define.xml", "03_metadata/define/define_sdtm.xml"):
        root = ET.parse(ROOT / relative).getroot()
        assert root.attrib[f"{{{ns['def']}}}Context"] == "Other"

    adam_root = ET.parse(ROOT / "03_metadata/define/define.xml").getroot()
    item = adam_root.find(".//odm:ItemDef[@OID='IT.ADAE.AESER']", ns)
    assert item is not None and item.attrib["Name"] == "AESER"
    assert adam_root.find(".//*[@OID='IT.ADAE.AESSER']") is None


def test_define_keys_and_source_sequences_match_authoritative_spec() -> None:
    ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
    root = ET.parse(ROOT / "03_metadata/define/define.xml").getroot()
    expected = {
        "ADSL": ["STUDYID", "USUBJID"],
        "ADEX": ["STUDYID", "USUBJID", "PARAMCD", "AVISIT"],
        "ADCM": ["STUDYID", "USUBJID", "CMSEQ"],
        "ADAE": ["STUDYID", "USUBJID", "AESEQ"],
        "ADLB": ["STUDYID", "USUBJID", "PARAMCD", "AVISITN", "LBDY", "LBSEQ"],
        "ADRS": ["STUDYID", "USUBJID", "PARAMCD", "ADT", "AVISIT"],
        "ADTTE": ["STUDYID", "USUBJID", "PARAMCD"],
    }
    item_names = {
        item.attrib["OID"]: item.attrib["Name"]
        for item in root.findall(".//odm:ItemDef", ns)
    }
    for group in root.findall(".//odm:ItemGroupDef", ns):
        refs = [
            (int(ref.attrib["KeySequence"]), item_names[ref.attrib["ItemOID"]])
            for ref in group.findall("odm:ItemRef", ns)
            if "KeySequence" in ref.attrib
        ]
        assert [name for _, name in sorted(refs)] == expected[group.attrib["Name"]]

    assert item_names["IT.ADCM.CMSEQ"] == "CMSEQ"
    assert item_names["IT.ADLB.LBSEQ"] == "LBSEQ"
    assert item_names["IT.ADRS.AVISIT"] == "AVISIT"
    assert "IT.ADRS.VISIT" not in item_names


def test_sdtm_define_declares_canonical_domain_keys() -> None:
    ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
    root = ET.parse(ROOT / "03_metadata/define/define_sdtm.xml").getroot()
    expected = {
        "DM": ["STUDYID", "USUBJID"],
        "AE": ["STUDYID", "USUBJID", "AESEQ"],
        "EX": ["STUDYID", "USUBJID", "EXSEQ"],
        "CM": ["STUDYID", "USUBJID", "CMSEQ"],
        "LB": ["STUDYID", "USUBJID", "LBSEQ"],
        "DS": ["STUDYID", "USUBJID", "DSSEQ"],
        "VS": ["STUDYID", "USUBJID", "VSSEQ"],
        "LS": ["STUDYID", "USUBJID", "LSSEQ"],
        "PN": ["STUDYID", "USUBJID", "PNSEQ"],
        "SUPPDM": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPAE": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPEX": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPCM": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPLB": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPDS": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "SUPPLS": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
        "TS": ["STUDYID", "TSSEQ"],
        "TA": ["STUDYID", "ARMCD", "TAETORD"],
    }
    item_names = {
        item.attrib["OID"]: item.attrib["Name"]
        for item in root.findall(".//odm:ItemDef", ns)
    }
    groups = root.findall(".//odm:ItemGroupDef", ns)
    assert {group.attrib["Name"] for group in groups} == set(expected)
    for group in groups:
        refs = [
            (int(ref.attrib["KeySequence"]), item_names[ref.attrib["ItemOID"]])
            for ref in group.findall("odm:ItemRef", ns)
            if "KeySequence" in ref.attrib
        ]
        assert [name for _, name in sorted(refs)] == expected[group.attrib["Name"]]


def test_workbook_matches_time_origin_and_no_imputation_contracts() -> None:
    workbook = load_workbook(
        ROOT / "03_metadata/adam/ADaM_spec.xlsx", read_only=True, data_only=False
    )
    variables = list(workbook["Variables"].iter_rows(values_only=True))
    headers = {name: idx for idx, name in enumerate(variables[0])}
    rows = {
        (row[headers["Dataset"]], row[headers["Variable"]]): row
        for row in variables[1:]
    }

    start = rows[("ADTTE", "STARTDT")]
    predecessor = str(start[headers["Predecessor"]])
    assert "RANDDT for OS, PFS, TTPSA, TTUMOR, and TTPAIN" in predecessor
    assert "TRTSDT for TTSAE" in predecessor

    for variable in ("ECOGBLIF", "PSABLIF", "ALPBLIF", "HGBBLIF", "ALBBLIF", "LDHBLIF"):
        row = rows[("ADSL", variable)]
        assert str(row[headers["Label"]]).endswith("Imputation Flag")

    methods = list(workbook["Methods"].iter_rows(values_only=True))
    method_headers = {name: idx for idx, name in enumerate(methods[0])}
    method_rows = {row[method_headers["ID"]]: row for row in methods[1:]}
    method = str(method_rows["MT.ADTTE_ADT"][method_headers["Description"]])
    assert "TTSAE starts at first dose" in method
    assert "may not precede STARTDT" in method
