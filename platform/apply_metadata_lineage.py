#!/usr/bin/env python3
"""Apply governed ADaM predecessor lineage to the authoring workbook."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path

import yaml
from lxml import etree
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC_PATH = ROOT / "03_metadata/adam" / "ADaM_spec.xlsx"
DEFINE_PATH = ROOT / "03_metadata/define" / "define.xml"
LINEAGE_PATH = ROOT / "config/metadata_lineage.yaml"
OUT_DIR = ROOT / "platform" / "metadata_lineage"
ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"
DEF_NS = "http://www.cdisc.org/ns/def/v2.1"
XML_NS = "http://www.w3.org/XML/1998/namespace"


def _load_lineage(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def _clean(value) -> str:
    return str(value or "").strip()


def _resolve_predecessor(dataset: str, variable: str, rules: dict) -> str:
    overrides = rules.get("overrides") or {}
    common = rules.get("common") or {}
    dataset_defaults = rules.get("dataset_defaults") or {}

    dataset_rules = overrides.get(dataset) or {}
    if variable in dataset_rules:
        return _clean(dataset_rules[variable])
    if variable in common:
        return _clean(common[variable])
    if dataset in dataset_defaults:
        return _clean(dataset_defaults[dataset])
    return ""


def _variable_rows(ws):
    headers = [_clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: pos + 1 for pos, name in enumerate(headers) if name}
    required = {"Dataset", "Variable", "Predecessor", "Method", "Origin"}
    missing = sorted(required - set(index))
    if missing:
        raise SystemExit(f"Variables sheet is missing required columns: {', '.join(missing)}")

    dataset_col = index["Dataset"]
    variable_col = index["Variable"]
    predecessor_col = index["Predecessor"]
    method_col = index["Method"]
    origin_col = index["Origin"]

    for row_num in range(2, ws.max_row + 1):
        dataset = _clean(ws.cell(row=row_num, column=dataset_col).value).upper()
        variable = _clean(ws.cell(row=row_num, column=variable_col).value).upper()
        if not dataset or not variable:
            continue
        current = _clean(ws.cell(row=row_num, column=predecessor_col).value)
        method = _clean(ws.cell(row=row_num, column=method_col).value)
        origin = _clean(ws.cell(row=row_num, column=origin_col).value)
        yield row_num, predecessor_col, method_col, origin_col, dataset, variable, current, method, origin


def _sheet_header_index(ws):
    headers = [_clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {name: pos + 1 for pos, name in enumerate(headers) if name}


def _method_rows(ws):
    index = _sheet_header_index(ws)
    required = {"ID", "Name", "Type", "Description"}
    missing = sorted(required - set(index))
    if missing:
        raise SystemExit(f"Methods sheet is missing required columns: {', '.join(missing)}")
    out = {}
    for row_num in range(2, ws.max_row + 1):
        method_id = _clean(ws.cell(row=row_num, column=index["ID"]).value)
        if not method_id:
            continue
        out[method_id] = {
            "row": row_num,
            "name": _clean(ws.cell(row=row_num, column=index["Name"]).value),
            "type": _clean(ws.cell(row=row_num, column=index["Type"]).value),
            "description": _clean(ws.cell(row=row_num, column=index["Description"]).value),
        }
    return out, index


def _define_method_state():
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(DEFINE_PATH), parser)
    root = tree.getroot()
    itemdefs = {
        node.get("OID"): node.get("Name")
        for node in root.xpath('//*[local-name()="ItemDef"]')
    }
    refs = {}
    for group in root.xpath('//*[local-name()="ItemGroupDef"]'):
        dataset = group.get("Name")
        if not dataset:
            continue
        for ref in group.xpath('./*[local-name()="ItemRef"]'):
            variable = itemdefs.get(ref.get("ItemOID"))
            if variable:
                refs[f"{dataset.upper()}.{variable.upper()}"] = ref
    methods = {
        node.get("OID"): node
        for node in root.xpath('//*[local-name()="MethodDef"]')
        if node.get("OID")
    }
    return tree, root, refs, methods


def _method_description(method_node):
    text = method_node.xpath('./*[local-name()="Description"]/*[local-name()="TranslatedText"]/text()')
    return _clean(text[0]) if text else ""


def _append_method_def(root, method_id, description):
    method = etree.Element(f"{{{ODM_NS}}}MethodDef")
    method.set("OID", method_id)
    method.set("Name", f"Method for {method_id.removeprefix('MT.')}")
    method.set("Type", "Computation")
    desc = etree.SubElement(method, f"{{{ODM_NS}}}Description")
    text = etree.SubElement(desc, f"{{{ODM_NS}}}TranslatedText")
    text.set(f"{{{XML_NS}}}lang", "en")
    text.text = description

    metadata = root.xpath('//*[local-name()="MetaDataVersion"]')[0]
    anchors = metadata.xpath('./*[local-name()="CommentDef"] | ./*[local-name()="AnalysisResultDisplays"]')
    if anchors:
        metadata.insert(metadata.index(anchors[0]), method)
    else:
        metadata.append(method)
    return method


def _write_define(tree):
    tree.write(
        str(DEFINE_PATH),
        encoding="UTF-8",
        xml_declaration=True,
        pretty_print=False,
    )


def apply_or_check(check_only: bool) -> dict:
    lineage = _load_lineage(LINEAGE_PATH)
    rules = lineage.get("predecessor_rules") or {}
    wb = load_workbook(SPEC_PATH)
    ws = wb["Variables"]
    method_ws = wb["Methods"]
    method_expansion = lineage.get("method_expansion") or {}
    method_assignments = {
        _clean(k).upper(): _clean(v)
        for k, v in (method_expansion.get("assignments") or {}).items()
    }
    origin_overrides = {
        _clean(k).upper(): _clean(v)
        for k, v in (lineage.get("origin_overrides") or {}).items()
    }
    method_definitions = {
        _clean(k): _clean(v)
        for k, v in (method_expansion.get("definitions") or {}).items()
    }
    undefined_assignments = sorted(set(method_assignments.values()) - set(method_definitions))
    if undefined_assignments:
        raise SystemExit(
            "Method expansion assignments are missing definitions: "
            + ", ".join(undefined_assignments)
        )
    method_rows, method_index = _method_rows(method_ws)
    define_tree, define_root, define_refs, define_methods = _define_method_state()

    rows = []
    missing_rules = []
    mismatches = []
    method_mismatches = []
    method_definition_mismatches = []
    define_method_mismatches = []
    origin_mismatches = []
    missing_define_refs = []
    applied = 0
    method_updates_applied = 0
    define_updates_applied = 0

    for row_num, predecessor_col, method_col, origin_col, dataset, variable, current, current_method, current_origin in _variable_rows(ws):
        expected = _resolve_predecessor(dataset, variable, rules)
        key = f"{dataset}.{variable}"
        if not expected:
            missing_rules.append(key)
            status = "MISSING_RULE"
        elif current != expected:
            mismatches.append(key)
            status = "MISMATCH"
            if not check_only:
                ws.cell(row=row_num, column=predecessor_col, value=expected)
                applied += 1
                status = "APPLIED"
        else:
            status = "MATCH"

        expected_method = method_assignments.get(key)
        method_status = "NOT_GOVERNED"
        if expected_method:
            if current_method != expected_method:
                method_mismatches.append(key)
                method_status = "METHOD_MISMATCH"
                if not check_only:
                    ws.cell(row=row_num, column=method_col, value=expected_method)
                    method_updates_applied += 1
                    method_status = "METHOD_APPLIED"
            else:
                method_status = "METHOD_MATCH"

            define_ref = define_refs.get(key)
            if define_ref is None:
                missing_define_refs.append(key)
            elif _clean(define_ref.get("MethodOID")) != expected_method:
                define_method_mismatches.append(key)
                if not check_only:
                    define_ref.set("MethodOID", expected_method)
                    define_updates_applied += 1

        expected_origin = origin_overrides.get(key)
        origin_status = "NOT_GOVERNED"
        if expected_origin:
            if current_origin != expected_origin:
                origin_mismatches.append(key)
                origin_status = "ORIGIN_MISMATCH"
                if not check_only:
                    ws.cell(row=row_num, column=origin_col, value=expected_origin)
                    method_updates_applied += 1
                    origin_status = "ORIGIN_APPLIED"
            else:
                origin_status = "ORIGIN_MATCH"

        rows.append({
            "dataset": dataset,
            "variable": variable,
            "current_predecessor": current,
            "expected_predecessor": expected,
            "current_method": current_method,
            "expected_method": expected_method or "",
            "current_origin": current_origin,
            "expected_origin": expected_origin or "",
            "status": status,
            "method_status": method_status,
            "origin_status": origin_status,
        })

    if missing_rules:
        raise SystemExit(
            "Metadata lineage rules do not cover all ADaM variables: "
            + ", ".join(missing_rules[:25])
            + (" ..." if len(missing_rules) > 25 else "")
        )

    for method_id, description in method_definitions.items():
        existing = method_rows.get(method_id)
        if existing is None:
            method_definition_mismatches.append(method_id)
            if not check_only:
                row_num = method_ws.max_row + 1
                method_ws.cell(row=row_num, column=method_index["ID"], value=method_id)
                method_ws.cell(row=row_num, column=method_index["Name"], value=f"Method for {method_id.removeprefix('MT.')}")
                method_ws.cell(row=row_num, column=method_index["Type"], value="Computation")
                method_ws.cell(row=row_num, column=method_index["Description"], value=description)
                method_updates_applied += 1
        elif existing["description"] != description or existing["type"] != "Computation":
            method_definition_mismatches.append(method_id)
            if not check_only:
                row_num = existing["row"]
                method_ws.cell(row=row_num, column=method_index["Name"], value=f"Method for {method_id.removeprefix('MT.')}")
                method_ws.cell(row=row_num, column=method_index["Type"], value="Computation")
                method_ws.cell(row=row_num, column=method_index["Description"], value=description)
                method_updates_applied += 1

        define_method = define_methods.get(method_id)
        if define_method is None:
            define_method_mismatches.append(method_id)
            if not check_only:
                _append_method_def(define_root, method_id, description)
                define_updates_applied += 1
        elif _method_description(define_method) != description:
            define_method_mismatches.append(method_id)
            if not check_only:
                text = define_method.xpath('./*[local-name()="Description"]/*[local-name()="TranslatedText"]')
                if text:
                    text[0].text = description
                else:
                    desc = etree.SubElement(define_method, f"{{{ODM_NS}}}Description")
                    translated = etree.SubElement(desc, f"{{{ODM_NS}}}TranslatedText")
                    translated.set(f"{{{XML_NS}}}lang", "en")
                    translated.text = description
                define_updates_applied += 1

    if missing_define_refs:
        raise SystemExit(
            "Define-XML is missing governed method assignment references: "
            + ", ".join(missing_define_refs)
        )

    total_mismatches = (
        len(mismatches)
        + len(method_mismatches)
        + len(method_definition_mismatches)
        + len(define_method_mismatches)
        + len(origin_mismatches)
    )

    if check_only and total_mismatches:
        status_value = "FAIL"
    else:
        status_value = "PASS"

    if not check_only and (applied or method_updates_applied):
        wb.save(SPEC_PATH)
    if not check_only and define_updates_applied:
        _write_define(define_tree)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUT_DIR / "metadata_lineage_application.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "dataset",
                "variable",
                "current_predecessor",
                "expected_predecessor",
                "current_method",
                "expected_method",
                "current_origin",
                "expected_origin",
                "status",
                "method_status",
                "origin_status",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    result = {
        "status": status_value,
        "mode": "check" if check_only else "apply",
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "lineage_file": str(LINEAGE_PATH.relative_to(ROOT)),
        "spec_file": str(SPEC_PATH.relative_to(ROOT)),
        "define_file": str(DEFINE_PATH.relative_to(ROOT)),
        "variables_assessed": len(rows),
        "mismatches": len(mismatches),
        "method_assignment_mismatches": len(method_mismatches),
        "method_definition_mismatches": len(method_definition_mismatches),
        "define_method_mismatches": len(define_method_mismatches),
        "origin_mismatches": len(origin_mismatches),
        "updates_applied": applied,
        "method_updates_applied": method_updates_applied,
        "define_updates_applied": define_updates_applied,
        "ct_dispositions": sorted((lineage.get("controlled_terminology_dispositions") or {}).keys()),
        "method_disposition_status": (lineage.get("method_dispositions") or {}).get("status", "missing"),
    }
    with (OUT_DIR / "metadata_lineage_status.json").open("w", encoding="utf-8") as handle:
        json.dump(result, handle, indent=2)

    if check_only and total_mismatches:
        print(f"Metadata lineage check failed: {total_mismatches} governed metadata mismatches")
        raise SystemExit(1)

    action = "checked" if check_only else "applied"
    print(
        f"Metadata lineage {action}: status={status_value}; "
        f"variables={len(rows)}; predecessor_mismatches={len(mismatches)}; "
        f"method_assignment_mismatches={len(method_mismatches)}; "
        f"method_definition_mismatches={len(method_definition_mismatches)}; "
        f"define_method_mismatches={len(define_method_mismatches)}; "
        f"origin_mismatches={len(origin_mismatches)}; "
        f"updates_applied={applied}; method_updates_applied={method_updates_applied}; "
        f"define_updates_applied={define_updates_applied}"
    )
    return result


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Apply or check governed ADaM metadata lineage")
    parser.add_argument("--check", action="store_true", help="Fail if workbook predecessors differ from config/metadata_lineage.yaml")
    args = parser.parse_args(argv)
    apply_or_check(check_only=args.check)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
