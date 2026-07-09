#!/usr/bin/env python3
"""Build a metadata governance report from spec, Define-XML, ARM, and conformance evidence."""

import argparse
import csv
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone

import yaml
from lxml import etree
from openpyxl import load_workbook


SPEC_PATH = "03_metadata/adam/ADaM_spec.xlsx"
DEFINE_ADAM_PATH = "03_metadata/define/define.xml"
DEFINE_SDTM_PATH = "03_metadata/define/define_sdtm.xml"
SPEC_DEFINE_PATH = "platform/conformance/spec_define_conformance.json"
SPEC_DATA_PATH = "platform/conformance/spec_data_conformance.json"
CT_CROSS_PATH = "platform/conformance/ct_cross_validation.json"
ADAM_CONF_STATUS_PATH = "06_qc_evidence/conformance/adam_conformance_status.json"
ADAM_CONF_REPORT_PATH = "06_qc_evidence/conformance/adam_conformance_report.csv"
TRACEABILITY_PATH = "06_qc_evidence/audit/adam_variable_traceability.csv"
METADATA_DRIFT_PATH = "06_qc_evidence/audit/metadata_data_drift.csv"
ARS_ARD_PATH = "05_outputs/ars/tropic_ard.csv"
ARS_EVENT_PATH = "05_outputs/ars/tropic_reporting_event.json"
METADATA_LINEAGE_PATH = "config/metadata_lineage.yaml"


def _load_json(path, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_yaml(path, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or default


def _read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _sheet_rows(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    out = []
    for values in rows[1:]:
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if any(v is not None and str(v).strip() != "" for v in row.values()):
            out.append(row)
    return out


def _define_counts(path):
    if not os.path.exists(path):
        return {"present": False}
    root = etree.parse(path).getroot()
    return {
        "present": True,
        "itemgroups": len(root.xpath('//*[local-name()="ItemGroupDef"]')),
        "items": len(root.xpath('//*[local-name()="ItemDef"]')),
        "codelists": len(root.xpath('//*[local-name()="CodeList"]')),
        "methods": len(root.xpath('//*[local-name()="MethodDef"]')),
        "valuelists": len(root.xpath('//*[local-name()="ValueListDef"]')),
        "resultdisplays": len(root.xpath('//*[local-name()="ResultDisplay"]')),
        "analysisresults": len(root.xpath('//*[local-name()="AnalysisResult"]')),
    }


def _spec_summary():
    datasets = _sheet_rows(SPEC_PATH, "Datasets")
    variables = _sheet_rows(SPEC_PATH, "Variables")
    valuelevel = _sheet_rows(SPEC_PATH, "ValueLevel")
    wheres = _sheet_rows(SPEC_PATH, "WhereClauses")
    codelists = _sheet_rows(SPEC_PATH, "Codelists")
    methods = _sheet_rows(SPEC_PATH, "Methods")
    dictionaries = _sheet_rows(SPEC_PATH, "Dictionaries")

    by_dataset = Counter(str(r.get("Dataset", "")).upper() for r in variables if r.get("Dataset"))
    origins = Counter(str(r.get("Origin", "") or "missing") for r in variables)
    missing_method_for_derived = [
        f"{str(r.get('Dataset')).upper()}.{str(r.get('Variable')).upper()}"
        for r in variables
        if str(r.get("Origin", "")).strip().lower() == "derived"
        and not str(r.get("Method", "") or "").strip()
    ]
    missing_predecessor = [
        f"{str(r.get('Dataset')).upper()}.{str(r.get('Variable')).upper()}"
        for r in variables
        if not str(r.get("Predecessor", "") or "").strip()
    ]
    blank_labels = [
        f"{str(r.get('Dataset')).upper()}.{str(r.get('Variable')).upper()}"
        for r in variables
        if not str(r.get("Label", "") or "").strip()
    ]
    return {
        "datasets": datasets,
        "variables": variables,
        "valuelevel": valuelevel,
        "wheres": wheres,
        "codelists": codelists,
        "methods": methods,
        "dictionaries": dictionaries,
        "by_dataset": by_dataset,
        "origins": origins,
        "missing_method_for_derived": missing_method_for_derived,
        "missing_predecessor": missing_predecessor,
        "blank_labels": blank_labels,
    }


def _dataset_rows(spec, spec_data):
    spec_data_by_ds = {}
    for item in (spec_data or {}).get("datasets", []):
        spec_data_by_ds[item.get("dataset", "").upper()] = item
    rows = []
    for ds in spec["datasets"]:
        name = str(ds.get("Dataset", "")).upper()
        data = spec_data_by_ds.get(name, {})
        rows.append({
            "dataset": name,
            "description": ds.get("Description", ""),
            "class": ds.get("Class", ""),
            "structure": ds.get("Structure", ""),
            "spec_variables": spec["by_dataset"].get(name, 0),
            "data_variables": data.get("n_data_vars", ""),
            "spec_data_status": data.get("status", "not_checked"),
            "missing_in_data": "|".join(data.get("missing_in_data", [])),
            "extra_in_data": "|".join(data.get("extra_in_data", [])),
            "ct_violations": data.get("ct_violations", ""),
            "type_mismatches": data.get("type_mismatches", ""),
            "length_mismatches": data.get("length_mismatches", ""),
        })
    return rows


def _write_csv(path, rows, fieldnames):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _md_table(headers, rows):
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        clean = [str(v).replace("|", "\\|").replace("\n", " ") for v in row]
        lines.append("| " + " | ".join(clean) + " |")
    return "\n".join(lines)


def _status_from_findings(findings):
    hard = [f for f in findings if f.get("severity") in {"major", "critical"}]
    if hard:
        return "warning"
    return "warning" if findings else "pass"


def build_metadata_control_report(out_dir, report_path):
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    spec = _spec_summary()
    adam_define = _define_counts(DEFINE_ADAM_PATH)
    sdtm_define = _define_counts(DEFINE_SDTM_PATH)
    spec_define = _load_json(SPEC_DEFINE_PATH, {})
    spec_data = _load_json(SPEC_DATA_PATH, {})
    ct_cross = _load_json(CT_CROSS_PATH, {})
    adam_conf = _load_json(ADAM_CONF_STATUS_PATH, {})
    adam_conf_rows = _read_csv(ADAM_CONF_REPORT_PATH)
    trace_rows = _read_csv(TRACEABILITY_PATH)
    drift_rows = _read_csv(METADATA_DRIFT_PATH)
    ars_rows = _read_csv(ARS_ARD_PATH)
    ars_event = _load_json(ARS_EVENT_PATH, {})
    metadata_lineage = _load_yaml(METADATA_LINEAGE_PATH, {})
    ct_dispositions = metadata_lineage.get("controlled_terminology_dispositions") or {}

    trace_status_counts = Counter(r.get("traceability_status", "missing") for r in trace_rows)
    drift_problem_rows = [
        r for r in drift_rows
        if r.get("missing_from_data") or r.get("not_in_define") or r.get("unlabelled_data_variables")
        or r.get("derived_variables_without_method")
    ]
    dataset_rows = _dataset_rows(spec, spec_data)
    conformance_rows = [
        ["Spec -> Define", spec_define.get("status", "missing"), spec_define.get("findings", "")],
        ["Spec -> Data", spec_data.get("status", "missing"), ""],
        ["CT cross-validation", (ct_cross.get("summary") or {}).get("status", "missing"), (ct_cross.get("summary") or {}).get("violations", "")],
        ["CT disposition register", "PASS" if ct_dispositions else "missing", f"{len(ct_dispositions)} sponsor-defined dispositions"],
        ["ADaM conformance status", adam_conf.get("status", "missing"), adam_conf.get("findings", "")],
    ]
    ct_gap_ids = [
        r.get("id")
        for r in ct_cross.get("codelists", [])
        if r.get("classification") == "unverifiable-numeric" or r.get("status") == "review"
    ]
    unresolved_ct_gap_ids = sorted(c for c in ct_gap_ids if c not in ct_dispositions)
    findings = []
    if spec_define.get("status") != "PASS":
        findings.append({"severity": "critical", "finding": "spec_define_conformance_not_pass", "detail": spec_define.get("status", "missing")})
    if spec_data.get("status") != "PASS":
        findings.append({"severity": "critical", "finding": "spec_data_conformance_not_pass", "detail": spec_data.get("status", "missing")})
    if (ct_cross.get("summary") or {}).get("status") == "SKIPPED":
        findings.append({"severity": "major", "finding": "ct_cross_validation_skipped", "detail": ct_cross.get("ct_source", "")})
    elif (ct_cross.get("summary") or {}).get("status") == "WARNING" and unresolved_ct_gap_ids:
        findings.append({
            "severity": "warning",
            "finding": "ct_cross_validation_unresolved_traceability_gaps",
            "detail": ", ".join(unresolved_ct_gap_ids),
        })
    if str(adam_conf.get("status", "")).upper().startswith("FAIL"):
        findings.append({"severity": "major", "finding": "adam_conformance_status_fail", "detail": f"{adam_conf.get('errors', '')} errors"})
    if spec["missing_predecessor"]:
        findings.append({"severity": "major", "finding": "spec_predecessor_not_documented", "detail": f"{len(spec['missing_predecessor'])} variables"})
    if drift_problem_rows:
        findings.append({"severity": "major", "finding": "metadata_data_drift_register_nonempty", "detail": f"{len(drift_problem_rows)} dataset rows"})
    if spec["blank_labels"]:
        findings.append({"severity": "major", "finding": "blank_spec_labels", "detail": f"{len(spec['blank_labels'])} variables"})

    status = {
        "status": _status_from_findings(findings),
        "generated_at": generated_at,
        "spec": {
            "datasets": len(spec["datasets"]),
            "variables": len(spec["variables"]),
            "valuelevel": len(spec["valuelevel"]),
            "whereclauses": len(spec["wheres"]),
            "codelist_terms": len(spec["codelists"]),
            "methods": len(spec["methods"]),
            "dictionaries": len(spec["dictionaries"]),
            "variables_missing_predecessor": len(spec["missing_predecessor"]),
            "derived_variables_missing_method": len(spec["missing_method_for_derived"]),
            "blank_labels": len(spec["blank_labels"]),
        },
        "define_adam": adam_define,
        "define_sdtm": sdtm_define,
        "conformance": {
            "spec_define_status": spec_define.get("status", "missing"),
            "spec_data_status": spec_data.get("status", "missing"),
            "ct_cross_validation_status": (ct_cross.get("summary") or {}).get("status", "missing"),
            "ct_disposition_count": len(ct_dispositions),
            "ct_unresolved_traceability_gaps": len(unresolved_ct_gap_ids),
            "adam_conformance_status": adam_conf.get("status", "missing"),
        },
        "ars": {
            "ard_rows": len(ars_rows),
            "reporting_event_present": bool(ars_event),
        },
        "findings": findings,
    }

    _write_csv(
        os.path.join(out_dir, "metadata_dataset_control.csv"),
        dataset_rows,
        [
            "dataset", "description", "class", "structure", "spec_variables",
            "data_variables", "spec_data_status", "missing_in_data", "extra_in_data",
            "ct_violations", "type_mismatches", "length_mismatches",
        ],
    )
    finding_rows = findings or [{"severity": "none", "finding": "none", "detail": ""}]
    _write_csv(os.path.join(out_dir, "metadata_findings.csv"), finding_rows, ["severity", "finding", "detail"])
    with open(os.path.join(out_dir, "metadata_control_status.json"), "w", encoding="utf-8") as f:
        json.dump(status, f, indent=2)

    lines = [
        "# TROPIC Metadata Control Report",
        "",
        f"Generated: {generated_at}",
        "",
        "> Metadata governance view across ADaM spec, Define-XML, ARM/ARS, conformance outputs, and traceability evidence. "
        "This report exposes both passing controls and unresolved metadata gaps.",
        "",
        "## Summary",
        "",
        _md_table(
            ["Item", "Value"],
            [
                ["Report status", status["status"]],
                ["Spec datasets", len(spec["datasets"])],
                ["Spec variables", len(spec["variables"])],
                ["Spec value-level rows", len(spec["valuelevel"])],
                ["Spec codelist terms", len(spec["codelists"])],
                ["Spec methods", len(spec["methods"])],
                ["ADaM Define ItemGroupDefs", adam_define.get("itemgroups", "")],
                ["ADaM Define ItemDefs", adam_define.get("items", "")],
                ["ADaM ARM ResultDisplays", adam_define.get("resultdisplays", "")],
                ["ADaM ARM AnalysisResults", adam_define.get("analysisresults", "")],
                ["SDTM Define ItemGroupDefs", sdtm_define.get("itemgroups", "")],
                ["ARS ARD rows", len(ars_rows)],
            ],
        ),
        "",
        "## Conformance Status",
        "",
        _md_table(["Control", "Status", "Findings/violations"], conformance_rows),
        "",
        "## Controlled Terminology Dispositions",
        "",
        _md_table(
            ["Codelist", "Status", "Applies to", "Rationale"],
            [
                [
                    codelist,
                    detail.get("status", ""),
                    ", ".join(detail.get("applies_to", [])),
                    detail.get("rationale", ""),
                ]
                for codelist, detail in sorted(ct_dispositions.items())
            ] or [["None", "", "", ""]],
        ),
        "",
        "## Dataset Metadata Control",
        "",
        _md_table(
            ["Dataset", "Class", "Spec vars", "Data vars", "Spec-data status", "CT violations", "Type mismatches", "Length mismatches"],
            [
                [r["dataset"], r["class"], r["spec_variables"], r["data_variables"], r["spec_data_status"], r["ct_violations"], r["type_mismatches"], r["length_mismatches"]]
                for r in dataset_rows
            ],
        ),
        "",
        "## Spec Variable Origin Profile",
        "",
        _md_table(["Origin", "Variables"], sorted(spec["origins"].items())),
        "",
        "## Traceability Extract Status",
        "",
        _md_table(["Traceability status", "Variables"], sorted(trace_status_counts.items())),
        "",
        "## Metadata Findings",
        "",
    ]
    if findings:
        lines.append(_md_table(["Severity", "Finding", "Detail"], [[f["severity"], f["finding"], f["detail"]] for f in findings]))
    else:
        lines.append("No metadata governance findings were detected.")

    if adam_conf_rows:
        lines.extend([
            "",
            "## ADaM Conformance Findings",
            "",
            _md_table(
                ["Rule", "Severity", "Dataset", "Variable", "Count", "Message"],
                [
                    [r.get("rule", ""), r.get("severity", ""), r.get("dataset", ""), r.get("variable", ""), r.get("count", ""), r.get("message", "")]
                    for r in adam_conf_rows[:50]
                ],
            ),
        ])

    if drift_problem_rows:
        lines.extend([
            "",
            "## Metadata/Data Drift Register Rows",
            "",
            _md_table(
                ["Standard", "Dataset", "Missing from data", "Not in define", "Unlabelled data vars", "Derived vars without method"],
                [
                    [
                        r.get("standard", ""), r.get("dataset", ""), r.get("missing_from_data", ""), r.get("not_in_define", ""),
                        r.get("unlabelled_data_variables", ""), r.get("derived_variables_without_method", ""),
                    ]
                    for r in drift_problem_rows
                ],
            ),
        ])

    lines.extend([
        "",
        "## Machine-Readable Outputs",
        "",
        "- `platform/metadata_control/metadata_control_status.json`",
        "- `platform/metadata_control/metadata_dataset_control.csv`",
        "- `platform/metadata_control/metadata_findings.csv`",
        "",
    ])
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return status


def main(argv=None):
    parser = argparse.ArgumentParser(description="Build metadata control report")
    parser.add_argument("--out-dir", default="platform/metadata_control")
    parser.add_argument("--report", default="docs/METADATA_CONTROL_REPORT.md")
    args = parser.parse_args(argv)

    status = build_metadata_control_report(args.out_dir, args.report)
    print(f"Metadata control status: {status['status']}")
    print(f"Wrote {args.report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
